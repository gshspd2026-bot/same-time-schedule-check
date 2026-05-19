from __future__ import annotations

import json
import re
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from typing import Any

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup


APP_TITLE = "동시간대 편성 체크"
PLAYWRIGHT_BROWSER_READY = False

# 크롤링 대상 사이트입니다.
# 사이트 구조가 바뀌면 fetch_tv_schedule(), fetch_homeshopping_schedule()만
# 먼저 확인하면 되도록 나머지 로직은 함수로 분리했습니다.
EPG_GUIDE_PAGE_URL = "http://www.epgguide.co.kr/page/tv.php"
EPG_GUIDE_PROGRAM_URL = "http://www.epgguide.co.kr/mod/ajax.get_program.php"
ECOMM_SCHEDULE_PAGE_URL = "https://live.ecomm-data.com/schedule/hs"
ECOMM_DATA_BASE_URL = "https://live.ecomm-data.com/_next/data"
HSMOA_SCHEDULE_URL = "https://api.hsmoa.net/v3/schedule"
IP_TV_GUIDE_URL = "http://211.43.210.44/tvguide/index.php"

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": ECOMM_SCHEDULE_PAGE_URL,
}

TV_REQUEST_HEADERS = {
    **REQUEST_HEADERS,
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Referer": EPG_GUIDE_PAGE_URL,
    "X-Requested-With": "XMLHttpRequest",
}


# EPG Guide 사이트에서 확인한 채널 코드입니다.
TV_CHANNEL_GROUPS = {
    "KBS1": [{"label": "KBS1", "category": "1", "media_code": "00002", "ip_main": "public", "ip_code": "9"}],
    "KBS2": [{"label": "KBS2", "category": "1", "media_code": "00003", "ip_main": "public", "ip_code": "7"}],
    "MBC": [{"label": "MBC", "category": "1", "media_code": "00004", "ip_main": "public", "ip_code": "11"}],
    "SBS": [{"label": "SBS", "category": "1", "media_code": "00005", "ip_main": "public", "ip_code": "6"}],
    "tvN": [{"label": "tvN", "category": "4", "media_code": "00230", "ip_main": "cable", "ip_code": "743"}],
    "종편": [
        {"label": "JTBC", "category": "13", "media_code": "00771", "ip_main": "organization", "ip_code": "570"},
        {"label": "TV조선", "category": "13", "media_code": "00773", "ip_main": "organization", "ip_code": "569"},
        {"label": "채널A", "category": "13", "media_code": "00772", "ip_main": "organization", "ip_code": "571"},
        {"label": "MBN", "category": "13", "media_code": "00770", "ip_main": "organization", "ip_code": "20"},
    ],
}

OFFICIAL_TV_SCHEDULE_URLS = {
    "KBS1": lambda target_date: (
        "https://schedule.kbs.co.kr/index.html"
        f"?sname=schedule&stype=table&type=globalList&search_day={target_date:%Y%m%d}&channel_group=G01"
    ),
    "KBS2": lambda target_date: (
        "https://schedule.kbs.co.kr/index.html"
        f"?sname=schedule&stype=table&type=globalList&search_day={target_date:%Y%m%d}&channel_group=G01"
    ),
    "MBC": lambda target_date: (
        f"https://schedule.imbc.com/?chcode=MBC&date={target_date:%Y%m%d}&m=0&c=0"
    ),
    "SBS": lambda target_date: (
        f"https://www.sbs.co.kr/schedule/index.html?type=tv&channel=SBS&pmDate={target_date:%Y%m%d}"
    ),
    "JTBC": lambda target_date: f"https://jtbc.co.kr/schedule/jtbc/{target_date:%Y%m%d}",
    "TV조선": lambda target_date: (
        f"https://broadcast.tvchosun.com/onair/schedule/today.cstv?date={target_date:%Y%m%d}"
    ),
    "채널A": lambda target_date: (
        f"https://ichannela.com/com/cmm/schedule.do?selectedDate={target_date:%Y%m%d}"
    ),
    "MBN": lambda target_date: f"https://www.mbn.co.kr/vod/schedule?date={target_date:%Y%m%d}",
}


# 라방바 데이터랩 홈쇼핑 편성표에서 사용하는 홈쇼핑 채널 코드입니다.
HOMESHOPPING_CHANNELS = {
    "CJ온스타일": "hs_cjonstyle",
    "롯데홈쇼핑": "hs_lotteimall",
    "현대홈쇼핑": "hs_hmall",
    "NS홈쇼핑": "hs_nsmall",
    "공영쇼핑": "hs_gongyoung",
    "홈앤쇼핑": "hs_hnsmall",
    "쇼핑엔티": "hs_shopntmall",
}

# 홈쇼핑모아 fallback API에서 사용하는 채널 코드입니다.
HSMOA_CHANNELS = {
    "CJ온스타일": "cjmall",
    "롯데홈쇼핑": "lotteimall",
    "현대홈쇼핑": "hmall",
    "NS홈쇼핑": "nsmall",
    "공영쇼핑": "immall",
    "홈앤쇼핑": "hnsmall",
    "쇼핑엔티": "shopnt",
}

HOMESHOPPING_CODE_TO_NAME = {
    **{v: k for k, v in HOMESHOPPING_CHANNELS.items()},
    **{v: k for k, v in HSMOA_CHANNELS.items()},
}

CORE_TV_CHANNEL_ORDER = {
    "KBS1": 0,
    "KBS2": 1,
    "MBC": 2,
    "SBS": 3,
}

HOMESHOPPING_CHANNEL_ORDER = {
    channel_name: index for index, channel_name in enumerate(HOMESHOPPING_CHANNELS)
}


TV_COLUMNS = [
    "채널",
    "프로그램명",
    "시작시간",
    "방송 예상 종료 시간",
    "내 방송 중 종료 여부",
]

HOMESHOPPING_COLUMNS = [
    "채널",
    "상품명",
    "분류",
    "방송 시작시간",
    "방송 종료시간",
]


def get_user_broadcast_window(
    broadcast_date: date,
    start_time_text: str,
    duration_minutes: int,
) -> tuple[datetime, datetime]:
    """입력한 방송날짜, 시작시간, 방송분으로 시작/종료 datetime을 계산합니다."""
    start_clock = parse_time_text(start_time_text)
    start_dt = datetime.combine(broadcast_date, start_clock)
    end_dt = start_dt + timedelta(minutes=int(duration_minutes))
    return start_dt, end_dt


def parse_time_text(value: str) -> time:
    """'21:00' 같은 문자열을 time 객체로 바꿉니다."""
    value = value.strip()
    if not re.match(r"^\d{1,2}:\d{2}$", value):
        raise ValueError("시간은 21:00 형식으로 입력해주세요.")

    hour_text, minute_text = value.split(":")
    hour = int(hour_text)
    minute = int(minute_text)

    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        raise ValueError("시간은 00:00부터 23:59 사이로 입력해주세요.")

    return time(hour=hour, minute=minute)


def resolve_tv_channels(selected_options: list[str]) -> list[dict[str, str]]:
    """화면에서 선택한 TV 옵션을 실제 조회할 채널 목록으로 펼칩니다."""
    channels: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()

    for option in selected_options:
        for channel in TV_CHANNEL_GROUPS.get(option, []):
            key = (channel["category"], channel["media_code"])
            if key not in seen:
                channels.append(channel)
                seen.add(key)

    return channels


def get_page(
    url: str,
    *,
    allow_insecure_retry: bool = False,
    session: requests.Session | None = None,
    **kwargs: Any,
) -> requests.Response:
    """
    웹 페이지/API를 가져옵니다.

    회사 보안망이나 VPN에서 자체 인증서를 끼워 넣으면 SSL 검증 오류가 날 수 있습니다.
    홈쇼핑 조회처럼 필요한 경우에만 allow_insecure_retry=True로 한 번 더 재시도합니다.
    """
    requester = session or requests
    try:
        return requester.get(url, **kwargs)
    except requests.exceptions.SSLError:
        if not allow_insecure_retry:
            raise

        requests.packages.urllib3.disable_warnings(
            requests.packages.urllib3.exceptions.InsecureRequestWarning
        )
        return requester.get(url, verify=False, **kwargs)


def get_json_response(response: requests.Response) -> dict[str, Any]:
    """응답이 JSON인지 확인한 뒤 dict로 변환합니다."""
    text = response.text.strip()
    if not text:
        raise ValueError("편성표 사이트가 빈 응답을 반환했습니다.")

    try:
        data = response.json()
    except ValueError as exc:
        preview = text[:80].replace("\n", " ")
        raise ValueError(f"JSON 응답이 아닙니다. 응답 앞부분: {preview}") from exc

    if not isinstance(data, dict):
        raise ValueError("편성표 사이트 응답 형식이 예상과 다릅니다.")

    return data


def extract_tv_html_from_response(response: requests.Response) -> str:
    """EPG Guide 응답에서 편성표 HTML을 꺼냅니다."""
    text = response.text.strip()
    if not text:
        raise ValueError("편성표 사이트가 빈 응답을 반환했습니다.")

    try:
        data = get_json_response(response)
        html = str(data.get("html", ""))
        if html:
            return html
    except ValueError:
        # 일부 환경에서는 JSON이 아니라 편성표 HTML 조각 또는 전체 HTML이 바로 올 수 있습니다.
        if "inner_dl" in text or "id=\"time" in text or "id='time" in text:
            return text

    raise ValueError("TV 편성표 사이트가 현재 실행 환경의 요청을 허용하지 않았습니다.")


def fetch_tv_schedule(
    selected_channels: list[dict[str, str]],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    선택한 TV 채널의 EPG Guide 편성표를 가져옵니다.

    EPG Guide는 JSON 응답 안에 HTML 테이블을 넣어 주므로, 그 HTML을 다시
    BeautifulSoup으로 파싱합니다.
    """
    errors: list[str] = []
    programs: list[dict[str, Any]] = []
    session = requests.Session()

    try:
        get_page(
            EPG_GUIDE_PAGE_URL,
            headers=TV_REQUEST_HEADERS,
            timeout=15,
            session=session,
        )
    except Exception:
        # 메인 페이지 사전 방문은 쿠키 확보용 보조 절차라 실패해도 실제 조회를 시도합니다.
        pass

    dates_to_fetch = [broadcast_start.date()]

    # 자정을 넘기는 방송은 다음날 00시 이후 편성도 필요합니다.
    # 밤 22시 이후 방송은 마지막 프로그램 종료시간 계산을 위해 다음날 첫 편성도 같이 봅니다.
    if broadcast_end.date() > broadcast_start.date() or broadcast_start.hour >= 22:
        dates_to_fetch.append(broadcast_start.date() + timedelta(days=1))

    for channel in selected_channels:
        for target_date in dates_to_fetch:
            ip_guide_programs = fetch_ip_tv_guide_schedule(channel, target_date)
            if ip_guide_programs:
                programs.extend(ip_guide_programs)
                continue

            official_programs = fetch_official_tv_schedule(
                channel["label"],
                target_date,
                session,
            )
            if official_programs:
                programs.extend(official_programs)
                continue

            params = {
                "cate_id": channel["category"],
                "media_code": channel["media_code"],
                "ymd": target_date.strftime("%Y%m%d"),
            }

            try:
                response = get_page(
                    EPG_GUIDE_PROGRAM_URL,
                    params=params,
                    headers=TV_REQUEST_HEADERS,
                    timeout=15,
                    session=session,
                )
                response.raise_for_status()
                html = extract_tv_html_from_response(response)
                programs.extend(parse_tv_schedule_html(html, channel["label"], target_date))
            except Exception as exc:
                playwright_programs = fetch_epg_schedule_with_playwright(
                    channel,
                    target_date,
                )
                if playwright_programs:
                    programs.extend(playwright_programs)
                else:
                    errors.append(f"{channel['label']}: {exc}")

    programs = add_tv_end_times(programs)
    if errors and not programs:
        return programs, [
            "TV 편성표 사이트가 현재 실행 환경의 요청을 허용하지 않아 TV 데이터를 불러오지 못했습니다. "
            "로컬 PC 실행에서는 정상 조회될 수 있습니다."
        ]
    if errors:
        return programs, ["일부 TV 채널 편성표를 불러오지 못했습니다."]

    return programs, errors


def fetch_ip_tv_guide_schedule(
    channel: dict[str, str],
    schedule_date: date,
) -> list[dict[str, Any]]:
    """IP 기반 TV Guide 페이지에서 선택 채널의 편성표를 가져옵니다."""
    ip_main = channel.get("ip_main")
    ip_code = channel.get("ip_code")
    if not ip_main or not ip_code:
        return []

    try:
        response = get_page(
            IP_TV_GUIDE_URL,
            params={
                "main": ip_main,
                "c": ip_code,
                "day": schedule_date.strftime("%Y_%m_%d"),
                "page": "",
            },
            headers={**REQUEST_HEADERS, "Referer": IP_TV_GUIDE_URL},
            timeout=15,
        )
        response.raise_for_status()
        response.encoding = "euc-kr"
        return parse_ip_tv_guide_html(response.text, channel["label"], schedule_date)
    except Exception:
        return []


def parse_ip_tv_guide_html(
    html: str,
    channel_name: str,
    schedule_date: date,
) -> list[dict[str, Any]]:
    """IP TV Guide HTML 테이블에서 선택 채널의 편성 시간과 프로그램명을 추출합니다."""
    soup = BeautifulSoup(html, "lxml")
    header_cells = soup.select("table#main_channel td")
    column_index = find_ip_tv_column_index(header_cells, channel_name, schedule_date)
    if column_index is None:
        return []

    rows = soup.select("table#result_tbl > tr")
    programs: list[dict[str, Any]] = []

    for row in rows:
        hour = parse_ip_tv_hour(row)
        if hour is None:
            continue

        channel_cells = row.find_all("td", recursive=False)
        # 첫 번째 td는 시간 컬럼이므로 데이터 컬럼 인덱스에 1을 더합니다.
        target_cell_index = column_index + 1
        if target_cell_index >= len(channel_cells):
            continue

        programs.extend(
            parse_ip_tv_program_cell(
                channel_cells[target_cell_index],
                channel_name,
                schedule_date,
                hour,
            )
        )

    return deduplicate_programs(programs)


def find_ip_tv_column_index(
    header_cells: list[Any],
    channel_name: str,
    schedule_date: date,
) -> int | None:
    """IP TV Guide의 헤더에서 선택할 데이터 컬럼 위치를 찾습니다."""
    labels: list[str] = []
    for cell in header_cells:
        text = normalize_space(cell.get_text(" ", strip=True))
        if not text or "시간" in text:
            continue
        if "◀" in text or "▶" in text:
            continue
        labels.append(text)

    # 채널을 지정하지 않은 목록형 페이지는 헤더가 KBS1/KBS2/MBC처럼 채널명입니다.
    for index, name in enumerate(labels):
        if name == channel_name:
            return index

    # c=채널코드로 들어간 상세 페이지는 헤더가 05.11/05.12처럼 날짜입니다.
    target_day = schedule_date.strftime("%m.%d")
    for index, name in enumerate(labels):
        if target_day in name:
            return index

    return None


def parse_ip_tv_hour(row: BeautifulSoup) -> int | None:
    """결과 테이블 행에서 시간 값을 읽습니다."""
    first_cell = row.find("td")
    if first_cell is None:
        return None

    text = normalize_space(first_cell.get_text(" ", strip=True))
    match = re.search(r"(\d{1,2})\s*시", text)
    if not match:
        return None

    hour = int(match.group(1))
    if 0 <= hour <= 23:
        return hour

    return None


def parse_ip_tv_program_cell(
    cell: BeautifulSoup,
    channel_name: str,
    schedule_date: date,
    hour: int,
) -> list[dict[str, Any]]:
    """한 채널/한 시간대 셀에서 여러 프로그램을 추출합니다."""
    programs: list[dict[str, Any]] = []
    rows = cell.select("table tr")

    for program_row in rows:
        cells = program_row.find_all("td")
        if len(cells) < 2:
            continue

        minute_text = normalize_space(cells[0].get_text(" ", strip=True))
        minute_match = re.search(r"(\d{2})", minute_text)
        if not minute_match:
            continue

        minute = int(minute_match.group(1))
        if minute > 59:
            continue

        title_cell = cells[1]
        for image in title_cell.find_all("img"):
            image.decompose()
        title = normalize_space(title_cell.get_text(" ", strip=True))
        if not title:
            continue

        programs.append(
            {
                "channel": channel_name,
                "program_name": title,
                "start": datetime.combine(schedule_date, time(hour=hour, minute=minute)),
            }
        )

    return programs


def fetch_epg_schedule_with_playwright(
    channel: dict[str, str],
    target_date: date,
) -> list[dict[str, Any]]:
    """requests가 막힐 때 Chromium 브라우저로 EPG Guide 편성표를 다시 시도합니다."""
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return []

    if not ensure_playwright_browser(sync_playwright):
        return []

    params = {
        "cate_id": channel["category"],
        "media_code": channel["media_code"],
        "ymd": target_date.strftime("%Y%m%d"),
    }
    query = "&".join(f"{key}={value}" for key, value in params.items())
    ajax_url = f"{EPG_GUIDE_PROGRAM_URL}?{query}"

    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"],
            )
            context = browser.new_context(
                user_agent=REQUEST_HEADERS["User-Agent"],
                locale="ko-KR",
                extra_http_headers={
                    "Accept-Language": REQUEST_HEADERS["Accept-Language"],
                    "Referer": EPG_GUIDE_PAGE_URL,
                    "X-Requested-With": "XMLHttpRequest",
                },
            )
            page = context.new_page()
            page.goto(EPG_GUIDE_PAGE_URL, wait_until="domcontentloaded", timeout=20000)
            page_html = load_epg_schedule_on_page(page, channel, target_date)
            if not page_html:
                page.goto(ajax_url, wait_until="domcontentloaded", timeout=20000)
                body_text = page.locator("body").inner_text(timeout=5000).strip()
                page_html = extract_tv_html_from_text(body_text) or page.content()
            context.close()
            browser.close()

        html = extract_tv_html_from_text(page_html)
        if not html:
            return []

        return parse_tv_schedule_html(html, channel["label"], target_date)
    except Exception:
        return []


def load_epg_schedule_on_page(page: Any, channel: dict[str, str], target_date: date) -> str:
    """EPG Guide의 실제 페이지에서 AJAX를 실행해 편성표 HTML을 가져옵니다."""
    try:
        result = page.evaluate(
            """
            async ({url, category, mediaCode, ymd}) => {
                const params = new URLSearchParams({
                    cate_id: category,
                    media_code: mediaCode,
                    ymd,
                });
                const response = await fetch(`${url}?${params.toString()}`, {
                    headers: {
                        "X-Requested-With": "XMLHttpRequest",
                        "Accept": "application/json, text/javascript, */*; q=0.01",
                    },
                    credentials: "same-origin",
                });
                const text = await response.text();
                try {
                    const data = JSON.parse(text);
                    return data.html || text;
                } catch (error) {
                    return text;
                }
            }
            """,
            {
                "url": EPG_GUIDE_PROGRAM_URL,
                "category": channel["category"],
                "mediaCode": channel["media_code"],
                "ymd": target_date.strftime("%Y%m%d"),
            },
        )
        return str(result or "")
    except Exception:
        return ""


def ensure_playwright_browser(sync_playwright: Any) -> bool:
    """Chromium 브라우저가 없으면 한 번 설치를 시도합니다."""
    global PLAYWRIGHT_BROWSER_READY
    if PLAYWRIGHT_BROWSER_READY:
        return True

    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"],
            )
            browser.close()
        PLAYWRIGHT_BROWSER_READY = True
        return True
    except Exception:
        pass

    try:
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            check=False,
            timeout=120,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"],
            )
            browser.close()
        PLAYWRIGHT_BROWSER_READY = True
        return True
    except Exception:
        return False


def extract_tv_html_from_text(text: str) -> str:
    """문자열에서 EPG Guide 편성표 HTML을 꺼냅니다."""
    text = text.strip()
    if not text:
        return ""

    try:
        data = json.loads(text)
        if isinstance(data, dict) and data.get("html"):
            return str(data["html"])
    except Exception:
        pass

    if "inner_dl" in text or "id=\"time" in text or "id='time" in text:
        return text

    return ""


def fetch_official_tv_schedule(
    channel_name: str,
    schedule_date: date,
    session: requests.Session,
) -> list[dict[str, Any]]:
    """방송사 공식 편성표 페이지에서 시간/프로그램명을 우선 가져옵니다."""
    url_builder = OFFICIAL_TV_SCHEDULE_URLS.get(channel_name)
    if url_builder is None:
        return []

    try:
        url = url_builder(schedule_date)
        response = get_page(
            url,
            headers={**REQUEST_HEADERS, "Referer": url},
            timeout=15,
            allow_insecure_retry=True,
            session=session,
        )
        response.raise_for_status()
        return parse_official_tv_schedule_html(response.text, channel_name, schedule_date)
    except Exception:
        return []


def parse_official_tv_schedule_html(
    html: str,
    channel_name: str,
    schedule_date: date,
) -> list[dict[str, Any]]:
    """
    공식 방송사 편성표 HTML에서 시간과 제목을 추출합니다.

    방송사마다 HTML 구조가 달라서 우선 텍스트 흐름에서 HH:MM 다음에 나오는
    프로그램명 후보를 찾습니다. 실패하면 기존 EPG Guide fallback이 동작합니다.
    """
    soup = BeautifulSoup(html, "lxml")
    selector_programs = parse_official_schedule_by_selectors(soup, channel_name, schedule_date)
    if len(selector_programs) >= 3:
        return selector_programs

    text_items = [normalize_space(text) for text in soup.stripped_strings]
    text_items = [text for text in text_items if text]

    programs: list[dict[str, Any]] = []
    seen_starts: set[datetime] = set()

    for index, text in enumerate(text_items):
        start_clock = parse_schedule_time_token(text)
        if start_clock is None:
            continue

        title = find_official_program_title(text_items, index + 1)
        if not title:
            continue

        start_dt = datetime.combine(schedule_date, start_clock)
        if start_dt in seen_starts:
            continue
        seen_starts.add(start_dt)

        programs.append(
            {
                "channel": channel_name,
                "program_name": title,
                "start": start_dt,
            }
        )

    if len(programs) < 3:
        return []

    return deduplicate_programs(programs)


def parse_official_schedule_by_selectors(
    soup: BeautifulSoup,
    channel_name: str,
    schedule_date: date,
) -> list[dict[str, Any]]:
    """방송사별로 알려진 시간/제목 클래스가 있으면 그 구조로 먼저 파싱합니다."""
    selector_pairs = [
        (".spthours", ".spititle"),  # SBS
        (".time", ".title"),
        (".hour", ".title"),
        (".program_time", ".program_title"),
        (".schedule_time", ".schedule_title"),
    ]

    for time_selector, title_selector in selector_pairs:
        time_nodes = soup.select(time_selector)
        title_nodes = soup.select(title_selector)
        if len(time_nodes) < 3 or len(title_nodes) < 3:
            continue

        programs: list[dict[str, Any]] = []
        for time_node, title_node in zip(time_nodes, title_nodes):
            start_clock = parse_schedule_time_token(time_node.get_text(strip=True))
            title = clean_official_program_title(title_node.get_text(" ", strip=True))
            if start_clock is None or not title:
                continue

            programs.append(
                {
                    "channel": channel_name,
                    "program_name": title,
                    "start": datetime.combine(schedule_date, start_clock),
                }
            )

        if len(programs) >= 3:
            return deduplicate_programs(programs)

    return []


def deduplicate_programs(programs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """같은 채널/시작시간/제목이 반복 파싱되면 한 번만 남깁니다."""
    results: list[dict[str, Any]] = []
    seen: set[tuple[str, datetime, str]] = set()
    for program in programs:
        key = (program["channel"], program["start"], program["program_name"])
        if key in seen:
            continue
        seen.add(key)
        results.append(program)
    return results


def parse_schedule_time_token(text: str) -> time | None:
    """편성표의 HH:MM 시간 토큰을 time 객체로 바꿉니다."""
    match = re.fullmatch(r"(\d{1,2}):(\d{2})", text.strip())
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2))
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None

    return time(hour=hour, minute=minute)


def find_official_program_title(text_items: list[str], start_index: int) -> str:
    """시간 다음에 나오는 텍스트 중 프로그램명으로 보이는 값을 고릅니다."""
    for text in text_items[start_index : start_index + 8]:
        if parse_schedule_time_token(text) is not None:
            return ""

        title = clean_official_program_title(text)
        if title:
            return title

    return ""


def clean_official_program_title(text: str) -> str:
    """공식 편성표 텍스트에서 등급/자막/다시보기 같은 부가 문구를 줄입니다."""
    text = normalize_space(text)
    if not text:
        return ""

    skip_words = {
        "편성표",
        "오늘",
        "이전",
        "다음",
        "다시보기",
        "인쇄하기",
        "NOW ON",
        "Now On",
        "ON AIR",
        "온에어",
        "달력보기",
        "날짜선택",
        "등록된 편성정보가 없습니다.",
    }
    if text in skip_words:
        return ""
    if len(text) <= 1:
        return ""
    if re.fullmatch(r"(All|A|HD|본|재|생|자막|수어|해설|ON|No-ON|[0-9]{1,2})[\s\w가-힣+-]*", text):
        return ""

    text = re.sub(r"^\[[^\]]+\]\s*", "", text)
    text = re.split(r"\s{2,}", text)[0]
    text = re.sub(
        r"\s+(All|A|HD|본|재|생|자막|수어|해설|ON|No-ON|폐쇄자막|화면해설|한국수어|[0-9]{1,2})"
        r"(\s+(All|A|HD|본|재|생|자막|수어|해설|ON|No-ON|폐쇄자막|화면해설|한국수어|[0-9]{1,2}))*$",
        "",
        text,
    )

    return text.strip()


def normalize_space(text: str) -> str:
    """여러 줄/공백 문자를 하나의 공백으로 정리합니다."""
    return " ".join(str(text).split())


def parse_tv_schedule_html(
    html: str,
    channel_name: str,
    schedule_date: date,
) -> list[dict[str, Any]]:
    """EPG Guide의 편성표 HTML에서 프로그램명과 시작시간을 뽑습니다."""
    soup = BeautifulSoup(html, "lxml")
    rows = soup.select("tr[id^='time']")
    programs: list[dict[str, Any]] = []

    for row in rows:
        hour = parse_hour_from_row(row.get("id", ""))
        if hour is None:
            continue

        for item in row.select("dl.inner_dl"):
            minute_tag = item.select_one("dt.tit")
            if minute_tag is None:
                continue

            minute_text = minute_tag.get_text(strip=True)
            if not minute_text.isdigit():
                continue

            minute = int(minute_text)
            # EPG Guide는 빈 칸을 61분처럼 표시하는 경우가 있어 실제 편성에서 제외합니다.
            if minute > 59:
                continue

            title = clean_tv_program_title(item)
            if not title:
                continue

            start_dt = datetime.combine(schedule_date, time(hour=hour, minute=minute))
            programs.append(
                {
                    "channel": channel_name,
                    "program_name": title,
                    "start": start_dt,
                }
            )

    return programs


def parse_hour_from_row(row_id: str) -> int | None:
    match = re.search(r"time(\d{2})", row_id)
    if not match:
        return None
    return int(match.group(1))


def clean_tv_program_title(item: BeautifulSoup) -> str:
    """프로그램명에서 재방송/자막 같은 아이콘 텍스트를 제거합니다."""
    content = item.select_one("dd.cont")
    if content is None:
        return ""

    for icon in content.select(".epgicon"):
        icon.decompose()

    return " ".join(content.get_text(" ", strip=True).split())


def add_tv_end_times(programs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """같은 채널의 다음 프로그램 시작시간에서 채널별 보정 시간을 뺀 값을 종료 예상 시간으로 사용합니다."""
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for program in programs:
        grouped[program["channel"]].append(program)

    results: list[dict[str, Any]] = []

    for channel_programs in grouped.values():
        channel_programs.sort(key=lambda item: item["start"])

        for index, program in enumerate(channel_programs):
            if index + 1 < len(channel_programs):
                next_start = channel_programs[index + 1]["start"]
                duration_minutes = int((next_start - program["start"]).total_seconds() // 60)
                end_dt = next_start - timedelta(
                    minutes=get_tv_end_offset_minutes(program["channel"], duration_minutes)
                )
            else:
                # 다음 편성을 알 수 없을 때의 보수적 fallback입니다.
                end_dt = program["start"] + timedelta(minutes=60)

            if end_dt < program["start"]:
                end_dt = program["start"]

            results.append({**program, "end": end_dt})

    results.sort(key=lambda item: (item["start"], item["channel"]))
    return results


def get_tv_end_offset_minutes(channel: str, duration_minutes: int) -> int:
    """채널/방송 길이에 따른 종료 예상 시간 보정값을 반환합니다."""
    if channel == "KBS1":
        return 3
    if channel in {"KBS2", "MBC", "SBS"}:
        return 8 if duration_minutes >= 60 else 7
    return 10


def filter_tv_schedule(
    programs: list[dict[str, Any]],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> pd.DataFrame:
    """
    내 방송 시간대와 관련 있는 TV 편성을 표 형태로 정리합니다.

    내 방송 중 종료 여부 조건:
    내 방송 시작시간 <= 공중파 프로그램 종료시간 <= 내 방송 종료시간
    """
    grouped_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for program in programs:
        program_start = program["start"]
        program_end = program["end"]

        # 내 방송 중에 종료되는 프로그램만 결과에 표시합니다.
        ends_during_my_broadcast = broadcast_start <= program_end <= broadcast_end
        if not ends_during_my_broadcast:
            continue

        grouped_rows[program["channel"]].append(
            {
                "채널": program["channel"],
                "프로그램명": program["program_name"],
                "시작시간": format_table_time(program_start, broadcast_start.date()),
                "방송 예상 종료 시간": format_table_time(program_end, broadcast_start.date()),
                "내 방송 중 종료 여부": "해당" if ends_during_my_broadcast else "비해당",
                "_start_dt": program_start,
                "_end_dt": program_end,
            }
        )

    rows: list[dict[str, Any]] = []
    for channel, channel_rows in grouped_rows.items():
        channel_rows.sort(key=lambda item: item["_end_dt"])
        rows.append(
            {
                "채널": channel,
                "프로그램명": join_table_values(item["프로그램명"] for item in channel_rows),
                "시작시간": join_table_values(item["시작시간"] for item in channel_rows),
                "방송 예상 종료 시간": join_table_values(
                    item["방송 예상 종료 시간"] for item in channel_rows
                ),
                "내 방송 중 종료 여부": join_table_values(
                    item["내 방송 중 종료 여부"] for item in channel_rows
                ),
                "_sort_end": min(item["_end_dt"] for item in channel_rows),
            }
        )

    rows.sort(key=lambda item: get_tv_row_sort_key(item))
    for row in rows:
        row.pop("_sort_end", None)

    return pd.DataFrame(rows, columns=TV_COLUMNS)


def join_table_values(values: Any) -> str:
    """같은 채널에 여러 편성이 있을 때 한 셀 안에 읽기 쉽게 묶습니다."""
    return " / ".join(str(value) for value in values if str(value).strip())


def get_tv_row_sort_key(row: dict[str, Any]) -> tuple[int, Any]:
    """KBS1, KBS2, MBC, SBS를 먼저 고정하고 나머지는 종료 예상 시간이 빠른 순서로 정렬합니다."""
    channel = str(row.get("채널", ""))
    if channel in CORE_TV_CHANNEL_ORDER:
        return (0, CORE_TV_CHANNEL_ORDER[channel])
    return (1, row.get("_sort_end", datetime.max))


def fetch_homeshopping_schedule(
    selected_channel_names: list[str],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    라방바 데이터랩 홈쇼핑 편성표에서 홈쇼핑 편성을 가져옵니다.

    생방 여부 확인 로직 추후 보완 필요:
    현재 페이지 데이터에서는 미래/과거 편성별 생방 여부를 안정적으로 구분할
    필드를 확인하지 못했습니다. 따라서 우선 선택한 시간대와 겹치는 홈쇼핑
    편성을 모두 보여줍니다.
    """
    errors: list[str] = []
    schedules: list[dict[str, Any]] = []

    if not selected_channel_names:
        return schedules, errors

    try:
        page_data = fetch_ecomm_page_data()
        build_id = str(page_data.get("buildId") or "")
        if not build_id:
            raise ValueError("Next.js buildId를 찾지 못했습니다.")
    except Exception as exc:
        return schedules, [f"홈쇼핑 편성표 페이지를 불러오지 못했습니다. ({exc})"]

    page_date_text = get_ecomm_page_date_text(page_data)
    page_schedule_items = extract_ecomm_schedule_items(page_data)

    for target_date in get_schedule_dates(broadcast_start, broadcast_end):
        target_date_text = target_date.strftime("%y%m%d")

        # 기본 페이지에 이미 선택 날짜의 편성표가 들어 있으면 그 데이터를 먼저 사용합니다.
        # JSON 데이터 URL이 빈 목록을 돌려주는 경우를 막기 위한 fallback입니다.
        if is_ecomm_page_data_for_target(page_date_text, target_date, page_schedule_items):
            schedules.extend(page_schedule_items)
            continue

        try:
            fetched_items = fetch_ecomm_schedule_for_date(build_id, target_date)
            if fetched_items:
                schedules.extend(fetched_items)
            elif is_ecomm_page_data_for_target(page_date_text, target_date, page_schedule_items):
                schedules.extend(page_schedule_items)
        except Exception as exc:
            # 선택 날짜가 오늘이고 JSON 데이터 요청이 실패하면 최초 페이지 안의 데이터를 한 번 더 사용합니다.
            if is_ecomm_page_data_for_target(page_date_text, target_date, page_schedule_items):
                schedules.extend(page_schedule_items)
            else:
                errors.append(
                    f"{target_date:%Y-%m-%d} 홈쇼핑 편성표를 불러오지 못했습니다. ({exc})"
                )

    if schedules:
        return schedules, errors

    fallback_schedules, fallback_errors = fetch_hsmoa_schedule(
        selected_channel_names,
        broadcast_start,
        broadcast_end,
    )
    if fallback_schedules:
        return fallback_schedules, []

    if errors:
        errors.extend(fallback_errors)
    else:
        errors.append(
            "홈쇼핑 편성표 데이터가 비어 있습니다. 선택한 날짜의 편성표가 아직 업데이트되지 않았을 수 있습니다."
        )
        errors.extend(fallback_errors)

    return schedules, errors


def fetch_hsmoa_schedule(
    selected_channel_names: list[str],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    라방바 조회가 막히거나 비어 있을 때 홈쇼핑모아 API에서 홈쇼핑 편성을 가져옵니다.

    이 fallback 단계에서는 분류 정보를 사용하지 않고, 채널/상품명/방송시간만 사용합니다.
    """
    if not selected_channel_names:
        return [], []

    base_dt = (broadcast_start - timedelta(hours=2)).replace(minute=0, second=0, microsecond=0)
    total_minutes = int((broadcast_end - base_dt).total_seconds() // 60)
    time_size = max(4, min(30, (total_minutes // 60) + 2))
    selected_hsmoa_codes = [HSMOA_CHANNELS[name] for name in selected_channel_names]
    params = {
        "base_hour_datetime": format_hsmoa_datetime(base_dt),
        "time_size": time_size,
        "direction": "down",
        "tv_channel": ",".join(selected_hsmoa_codes),
    }

    try:
        response = get_page(
            HSMOA_SCHEDULE_URL,
            params=params,
            headers=REQUEST_HEADERS,
            timeout=20,
            allow_insecure_retry=True,
        )
        response.raise_for_status()
        response.encoding = "utf-8"
        return extract_hsmoa_schedule_items(response.json()), []
    except Exception as exc:
        return [], [f"홈쇼핑모아 편성표를 불러오지 못했습니다. ({exc})"]


def format_hsmoa_datetime(value: datetime) -> str:
    """홈쇼핑모아 API 요청에 사용할 한국시간 ISO 문자열을 만듭니다."""
    return value.strftime("%Y-%m-%dT%H:%M:%S+09:00")


def extract_hsmoa_schedule_items(data: Any) -> list[dict[str, Any]]:
    """홈쇼핑모아 schedule 응답에서 실제 상품 편성 목록을 평평하게 펼칩니다."""
    if not isinstance(data, dict):
        return []

    items: list[dict[str, Any]] = []
    for section_name in ("before_live", "live", "after_live", "future", "past"):
        section = data.get(section_name)
        if not isinstance(section, list):
            continue

        for block in section:
            if not isinstance(block, dict):
                continue

            candidates = block.get("schedules") or block.get("products") or []
            if not isinstance(candidates, list):
                continue

            for item in candidates:
                if isinstance(item, dict):
                    items.append({**item, "_source": "hsmoa"})

    return items


def fetch_ecomm_page_data() -> dict[str, Any]:
    """홈쇼핑 편성표 페이지의 __NEXT_DATA__ JSON을 읽습니다."""
    response = get_page(
        ECOMM_SCHEDULE_PAGE_URL,
        headers=REQUEST_HEADERS,
        timeout=20,
        allow_insecure_retry=True,
    )
    response.raise_for_status()
    response.encoding = "utf-8"
    return extract_next_data(response.text)


def extract_next_data(html: str) -> dict[str, Any]:
    """Next.js 페이지 HTML에서 __NEXT_DATA__ 스크립트 내용을 파싱합니다."""
    soup = BeautifulSoup(html, "lxml")
    script = soup.find("script", id="__NEXT_DATA__")
    if script is None:
        raise ValueError("__NEXT_DATA__ 스크립트를 찾지 못했습니다.")

    script_text = script.string or script.get_text(strip=True)
    if not script_text:
        raise ValueError("__NEXT_DATA__ 내용이 비어 있습니다.")

    parsed = json.loads(script_text)
    if not isinstance(parsed, dict):
        raise ValueError("__NEXT_DATA__ 형식이 올바르지 않습니다.")
    return parsed


def get_ecomm_page_date_text(data: dict[str, Any]) -> str:
    """라방바 데이터랩 페이지가 들고 있는 조회 날짜를 YYMMDD 형식으로 꺼냅니다."""
    page_props = data.get("pageProps")
    if not isinstance(page_props, dict):
        return ""

    page_date = page_props.get("d")
    return str(page_date or "").strip()


def is_ecomm_page_data_for_target(
    page_date_text: str,
    target_date: date,
    page_schedule_items: list[dict[str, Any]],
) -> bool:
    """기본 페이지 안의 홈쇼핑 데이터가 사용자가 조회한 날짜와 맞는지 확인합니다."""
    if not page_schedule_items:
        return False

    target_date_text = target_date.strftime("%y%m%d")
    if page_date_text == target_date_text:
        return True

    # 사이트가 d 값을 비워도 기본 페이지가 당일 편성을 들고 있는 경우가 있습니다.
    return not page_date_text and target_date == date.today()


def fetch_ecomm_schedule_for_date(build_id: str, target_date: date) -> list[dict[str, Any]]:
    """Next.js 데이터 JSON에서 특정 날짜의 홈쇼핑 편성 목록을 가져옵니다."""
    response = get_page(
        f"{ECOMM_DATA_BASE_URL}/{build_id}/schedule/hs.json",
        params={"date": target_date.strftime("%y%m%d")},
        headers=REQUEST_HEADERS,
        timeout=20,
        allow_insecure_retry=True,
    )
    response.raise_for_status()
    response.encoding = "utf-8"
    payload = response.json()
    return extract_ecomm_schedule_items(payload)


def extract_ecomm_schedule_items(data: Any) -> list[dict[str, Any]]:
    """라방바 데이터랩 응답에서 실제 홈쇼핑 편성 list만 꺼냅니다."""
    if not isinstance(data, dict):
        return []

    page_props = data.get("pageProps")
    if not isinstance(page_props, dict):
        return []

    items = page_props.get("list")
    if not isinstance(items, list):
        return []

    return [item for item in items if isinstance(item, dict)]


def get_schedule_dates(broadcast_start: datetime, broadcast_end: datetime) -> list[date]:
    """자정을 넘기는 방송을 위해 조회해야 할 날짜 목록을 만듭니다."""
    dates: list[date] = []
    current_date = broadcast_start.date()
    last_date = broadcast_end.date()

    while current_date <= last_date:
        dates.append(current_date)
        current_date += timedelta(days=1)

    return dates


def filter_homeshopping_schedule(
    schedules: list[dict[str, Any]],
    selected_channel_names: list[str],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> pd.DataFrame:
    """
    내 방송 시간대와 겹치는 홈쇼핑 편성을 추출합니다.

    조건:
    홈쇼핑 방송 시작시간 < 내 방송 종료시간
    AND 홈쇼핑 방송 종료시간 > 내 방송 시작시간
    """
    selected_codes = get_selected_homeshopping_codes(selected_channel_names)
    rows: list[dict[str, Any]] = []
    seen_keys: set[str] = set()

    for item in schedules:
        channel_code = str(
            item.get("platform_id")
            or item.get("tv_channel")
            or item.get("site")
            or item.get("channel")
            or item.get("channel_code")
            or ""
        )
        channel_name = str(item.get("platform_name") or "").strip()

        if selected_codes and channel_code and channel_code not in selected_codes:
            continue
        if selected_codes and not channel_code and channel_name not in selected_channel_names:
            continue

        start_dt = parse_datetime_value(
            item.get("hsshow_datetime_start")
            or item.get("start_datetime")
            or item.get("broadcast_start_datetime")
            or item.get("startDateTime")
            or item.get("start_time")
        )
        end_dt = parse_datetime_value(
            item.get("hsshow_datetime_end")
            or item.get("end_datetime")
            or item.get("broadcast_end_datetime")
            or item.get("endDateTime")
            or item.get("end_time")
        )

        if start_dt is None or end_dt is None:
            continue

        if not (start_dt < broadcast_end and end_dt > broadcast_start):
            continue

        product_name = str(
            item.get("hsshow_title")
            or item.get("name")
            or item.get("product_name")
            or item.get("productName")
            or item.get("title")
            or ""
        ).strip()
        is_hsmoa_fallback = item.get("_source") == "hsmoa"
        category_name = "" if is_hsmoa_fallback else get_homeshopping_category_name(item)
        unique_key = (
            f"{channel_code}|{product_name}|{start_dt:%Y%m%d%H%M}|{end_dt:%Y%m%d%H%M}"
        )
        if unique_key in seen_keys:
            continue
        seen_keys.add(unique_key)

        display_channel = HOMESHOPPING_CODE_TO_NAME.get(
            channel_code,
            channel_name or channel_code or "확인 필요",
        )
        rows.append(
            {
                "채널": display_channel,
                "상품명": product_name or "상품명 확인 필요",
                "분류": category_name or ("" if is_hsmoa_fallback else "분류 확인 필요"),
                "방송 시작시간": format_table_time(start_dt, broadcast_start.date()),
                "방송 종료시간": format_table_time(end_dt, broadcast_start.date()),
                "_sort_start": start_dt,
            }
        )

    rows.sort(key=lambda row: get_homeshopping_row_sort_key(row))
    for row in rows:
        row.pop("_sort_start", None)

    return pd.DataFrame(rows, columns=HOMESHOPPING_COLUMNS)


def get_selected_homeshopping_codes(selected_channel_names: list[str]) -> set[str]:
    """라방바 코드와 홈쇼핑모아 fallback 코드를 함께 선택 코드로 사용합니다."""
    selected_codes: set[str] = set()
    for name in selected_channel_names:
        if name in HOMESHOPPING_CHANNELS:
            selected_codes.add(HOMESHOPPING_CHANNELS[name])
        if name in HSMOA_CHANNELS:
            selected_codes.add(HSMOA_CHANNELS[name])
    return selected_codes


def get_homeshopping_row_sort_key(row: dict[str, Any]) -> tuple[int, int, Any]:
    """홈쇼핑 채널을 CJ, 롯데, 현대, NS, 공영, 홈앤, 쇼핑엔티 순서로 정렬합니다."""
    channel = str(row.get("채널", ""))
    channel_order = HOMESHOPPING_CHANNEL_ORDER.get(channel)
    if channel_order is not None:
        return (0, channel_order, row.get("_sort_start", datetime.max))
    return (1, 999, row.get("_sort_start", datetime.max))


def get_homeshopping_category_name(item: dict[str, Any]) -> str:
    """홈쇼핑 편성의 분류명을 꺼냅니다."""
    category = item.get("cat")
    if isinstance(category, dict):
        category_name = category.get("cat_name") or category.get("name")
        if category_name:
            return str(category_name).strip()

    return str(item.get("cat_name") or item.get("category") or "").strip()


def parse_datetime_value(value: Any) -> datetime | None:
    """API 응답의 여러 datetime 표현을 Python datetime으로 바꿉니다."""
    if value is None:
        return None

    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value).strip()
        if not text:
            return None

        # ISO 문자열이 아닌 경우를 대비한 보조 파서입니다.
        candidates = [
            text.replace("Z", "+00:00"),
            text.replace("/", "-"),
        ]
        parsed = None
        for candidate in candidates:
            try:
                parsed = datetime.fromisoformat(candidate)
                break
            except ValueError:
                continue

        if parsed is None:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y%m%d%H%M"):
                try:
                    parsed = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue

        if parsed is None:
            return None

    if parsed.tzinfo is not None:
        korea_tz = timezone(timedelta(hours=9))
        parsed = parsed.astimezone(korea_tz).replace(tzinfo=None)

    return parsed


def get_sample_tv_schedule(
    selected_channels: list[dict[str, str]],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> list[dict[str, Any]]:
    """크롤링 실패 시 앱 화면과 필터링 로직을 확인하기 위한 샘플 TV 편성입니다."""
    samples: list[dict[str, Any]] = []

    for index, channel in enumerate(selected_channels[:8]):
        first_start = broadcast_start - timedelta(minutes=40 + index * 3)
        first_end = broadcast_start + timedelta(minutes=15 + index * 5)
        second_end = min(broadcast_end + timedelta(minutes=20), first_end + timedelta(minutes=50))

        samples.append(
            {
                "channel": channel["label"],
                "program_name": f"샘플 프로그램 {index + 1}",
                "start": first_start,
                "end": first_end,
            }
        )
        samples.append(
            {
                "channel": channel["label"],
                "program_name": f"샘플 뉴스/예능 {index + 1}",
                "start": first_end,
                "end": second_end,
            }
        )

    return samples


def get_sample_homeshopping_schedule(
    selected_channel_names: list[str],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> list[dict[str, Any]]:
    """크롤링 실패 시 앱 화면과 필터링 로직을 확인하기 위한 샘플 홈쇼핑 편성입니다."""
    samples: list[dict[str, Any]] = []

    for index, channel_name in enumerate(selected_channel_names[:8]):
        start_dt = broadcast_start - timedelta(minutes=20 - index * 5)
        end_dt = min(broadcast_end + timedelta(minutes=15), start_dt + timedelta(minutes=65))

        samples.append(
            {
                "platform_id": HOMESHOPPING_CHANNELS[channel_name],
                "platform_name": channel_name,
                "hsshow_title": f"샘플 상품 {index + 1}",
                "hsshow_datetime_start": start_dt.strftime("%Y%m%d%H%M"),
                "hsshow_datetime_end": end_dt.strftime("%Y%m%d%H%M"),
                "cat": {"cat_name": "샘플 분류"},
            }
        )

    return samples


def format_table_time(value: datetime, reference_date: date) -> str:
    """같은 날짜는 HH:MM, 날짜가 다르면 MM-DD HH:MM으로 표시합니다."""
    if value.date() == reference_date:
        return value.strftime("%H:%M")
    return value.strftime("%m-%d %H:%M")


def format_broadcast_window(start_dt: datetime, end_dt: datetime) -> str:
    if start_dt.date() != end_dt.date():
        return f"{start_dt:%Y-%m-%d %H:%M}~{end_dt:%Y-%m-%d %H:%M}"
    return f"{start_dt:%Y-%m-%d %H:%M}~{end_dt:%H:%M}"


def export_results(
    broadcast_start: datetime,
    broadcast_end: datetime,
    tv_df: pd.DataFrame,
    homeshopping_df: pd.DataFrame,
    tv_programs: list[dict[str, Any]],
) -> bytes:
    """결과를 엑셀 파일로 만들고, TV 종료 예상 시간을 도식화합니다."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "편성결과"
        summary_df = pd.DataFrame(
            [{"항목": "내 방송 시간", "내용": format_broadcast_window(broadcast_start, broadcast_end)}]
        )
        summary_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0)
        worksheet = writer.sheets[sheet_name]

        row = len(summary_df) + 3
        worksheet.cell(row=row, column=1, value="공중파/TV 편성표")
        tv_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=row)

        row += len(tv_df) + 3
        worksheet.cell(row=row, column=1, value="홈쇼핑 편성표")
        homeshopping_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=row)

        adjust_excel_sheet_widths(worksheet)
        render_tv_ending_diagram_sheet(
            writer.book,
            tv_programs,
            broadcast_start,
            broadcast_end,
        )

    return output.getvalue()


def render_tv_ending_diagram_sheet(
    workbook: Any,
    tv_programs: list[dict[str, Any]],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> None:
    """방송 종료 예상 시간을 1분 단위 시간축 위 5분짜리 블럭으로 도식화합니다."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    sheet_name = "방송종료도식"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name, 0)

    total_minutes = max(1, int((broadcast_end - broadcast_start).total_seconds() // 60))
    max_minutes = min(total_minutes, 600)
    block_programs = get_tv_ending_block_programs(tv_programs, broadcast_start, broadcast_end)

    worksheet["A1"] = "동시간대 편성 체크"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = f"내 방송 시간: {format_broadcast_window(broadcast_start, broadcast_end)}"
    worksheet["A4"] = "채널"
    worksheet["A6"] = "시"
    worksheet["A7"] = "분"

    header_fill = PatternFill("solid", fgColor="E6E6E6")
    minute_fill = PatternFill("solid", fgColor="F4F4F4")
    thin_gray = Side(style="thin", color="D9D9D9")
    medium_dark = Side(style="medium", color="404040")

    for col_offset in range(max_minutes):
        current = broadcast_start + timedelta(minutes=col_offset)
        col = 2 + col_offset
        hour_cell = worksheet.cell(row=6, column=col, value=current.hour)
        minute_cell = worksheet.cell(row=7, column=col, value=current.minute)
        hour_cell.fill = header_fill
        minute_cell.fill = minute_fill
        hour_cell.alignment = Alignment(horizontal="center", vertical="center")
        minute_cell.alignment = Alignment(horizontal="center", vertical="center")
        hour_cell.border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)
        minute_cell.border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)
        worksheet.column_dimensions[get_column_letter(col)].width = 3

    row = 8
    for program in block_programs:
        end_dt = program["end"]
        minute_offset = int((end_dt - broadcast_start).total_seconds() // 60)
        if minute_offset < 0 or minute_offset >= max_minutes:
            continue

        start_col = 2 + minute_offset
        end_col = min(start_col + 4, 1 + max_minutes)
        channel = str(program["channel"])
        label = f"{channel} {program['program_name']} {format_table_time(end_dt, broadcast_start.date())}"

        worksheet.cell(row=row, column=1, value=channel)
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        worksheet.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        block_cell = worksheet.cell(row=row, column=start_col, value=label)
        block_cell.fill = PatternFill("solid", fgColor=get_tv_block_color(channel))
        block_cell.font = Font(bold=True, size=8)
        block_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=col).border = Border(
                top=medium_dark,
                bottom=medium_dark,
                left=medium_dark if col == start_col else thin_gray,
                right=medium_dark if col == end_col else thin_gray,
            )

        worksheet.row_dimensions[row].height = 28
        row += 1

    if not block_programs:
        worksheet["A8"] = "내 방송 중 종료되는 TV 편성이 없습니다."

    worksheet.column_dimensions["A"].width = 12
    worksheet.freeze_panes = "B8"


def get_tv_ending_block_programs(
    tv_programs: list[dict[str, Any]],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> list[dict[str, Any]]:
    """내 방송 중 종료되는 TV 프로그램만 종료 시간순으로 정리합니다."""
    programs = [
        program
        for program in tv_programs
        if broadcast_start <= program.get("end", datetime.min) <= broadcast_end
    ]
    programs.sort(
        key=lambda item: (
            item["end"],
            CORE_TV_CHANNEL_ORDER.get(str(item["channel"]), 99),
            str(item["channel"]),
        )
    )
    return programs


def get_tv_block_color(channel: str) -> str:
    """종료 도식 블럭의 채널별 배경색입니다."""
    colors = {
        "KBS1": "D9EAF7",
        "KBS2": "B7DEE8",
        "MBC": "D8EAD2",
        "SBS": "FCE4D6",
        "tvN": "E4DFEC",
        "JTBC": "FFF2CC",
        "TV조선": "EADCF8",
        "채널A": "DDEBF7",
        "MBN": "E2F0D9",
    }
    return colors.get(channel, "E7E6E6")


def adjust_excel_sheet_widths(worksheet: Any) -> None:
    """엑셀 한 시트 안의 열너비를 내용에 맞춰 보기 좋게 조정합니다."""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    for row in worksheet.iter_rows():
        first_cell_value = row[0].value if row else None
        if first_cell_value in ("공중파/TV 편성표", "홈쇼핑 편성표"):
            row[0].font = Font(bold=True)

        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            longest_line = max((len(line) for line in value.splitlines()), default=0)
            max_length = max(max_length, longest_line)

        adjusted_width = min(max(max_length + 2, 10), 55)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    worksheet.freeze_panes = "A2"


def build_clipboard_text(
    broadcast_start: datetime,
    broadcast_end: datetime,
    tv_df: pd.DataFrame,
    homeshopping_df: pd.DataFrame,
) -> str:
    """엑셀/스프레드시트에 바로 붙여넣기 좋은 탭 구분 텍스트를 만듭니다."""
    sections = [
        f"내 방송 시간\t{format_broadcast_window(broadcast_start, broadcast_end)}",
        "",
        "[공중파/TV 편성표]",
        tv_df.to_csv(index=False, sep="\t", lineterminator="\n").strip(),
        "",
        "[홈쇼핑 편성표]",
        homeshopping_df.to_csv(index=False, sep="\t", lineterminator="\n").strip(),
    ]
    return "\n".join(sections)


def render_results(
    broadcast_start: datetime,
    broadcast_end: datetime,
    tv_df: pd.DataFrame,
    homeshopping_df: pd.DataFrame,
    tv_programs: list[dict[str, Any]],
    tv_errors: list[str],
    homeshopping_errors: list[str],
) -> None:
    """Streamlit 화면에 요약, 표, 다운로드 버튼을 출력합니다."""
    st.subheader("조회 결과")
    st.info(f"내 방송 시간: {format_broadcast_window(broadcast_start, broadcast_end)}")

    if tv_errors or homeshopping_errors:
        st.warning("데이터를 불러오지 못했습니다. 일부 결과가 샘플이거나 표시되지 않을 수 있습니다.")
        for message in tv_errors + homeshopping_errors:
            st.caption(message)

    st.subheader("공중파/TV 편성표")
    st.dataframe(tv_df, use_container_width=True, hide_index=True)

    st.subheader("홈쇼핑 편성표")
    if homeshopping_df.empty and not homeshopping_errors:
        st.info("선택한 방송 시간대와 겹치는 홈쇼핑 편성이 없습니다. 날짜, 시간, 채널 선택을 확인해주세요.")
    st.dataframe(homeshopping_df, use_container_width=True, hide_index=True)

    clipboard_text = build_clipboard_text(
        broadcast_start,
        broadcast_end,
        tv_df,
        homeshopping_df,
    )

    st.subheader("복사")
    st.caption("아래 칸을 클릭한 뒤 Ctrl+A, Ctrl+C로 복사하세요.")
    st.text_area("클립보드 복사용 텍스트", value=clipboard_text, height=260)

    try:
        excel_bytes = export_results(
            broadcast_start,
            broadcast_end,
            tv_df,
            homeshopping_df,
            tv_programs,
        )
        st.download_button(
            label="엑셀 다운로드",
            data=excel_bytes,
            file_name=f"동시간대_편성_체크_{broadcast_start:%Y%m%d_%H%M}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception:
        st.download_button(
            label="CSV 다운로드",
            data=clipboard_text.encode("utf-8-sig"),
            file_name=f"동시간대_편성_체크_{broadcast_start:%Y%m%d_%H%M}.csv",
            mime="text/csv",
        )


def run_search(
    selected_tv_options: list[str],
    selected_homeshopping_channels: list[str],
    broadcast_start: datetime,
    broadcast_end: datetime,
) -> tuple[pd.DataFrame, pd.DataFrame, list[dict[str, Any]], list[str], list[str]]:
    """조회, fallback, 필터링을 한 번에 실행합니다."""
    tv_channels = resolve_tv_channels(selected_tv_options)

    tv_programs, tv_errors = fetch_tv_schedule(tv_channels, broadcast_start, broadcast_end)
    if tv_errors and not tv_programs:
        tv_programs = get_sample_tv_schedule(tv_channels, broadcast_start, broadcast_end)

    homeshopping_schedules, homeshopping_errors = fetch_homeshopping_schedule(
        selected_homeshopping_channels,
        broadcast_start,
        broadcast_end,
    )

    tv_df = filter_tv_schedule(tv_programs, broadcast_start, broadcast_end)
    homeshopping_df = filter_homeshopping_schedule(
        homeshopping_schedules,
        selected_homeshopping_channels,
        broadcast_start,
        broadcast_end,
    )
    if homeshopping_df.empty:
        fallback_schedules, fallback_errors = fetch_hsmoa_schedule(
            selected_homeshopping_channels,
            broadcast_start,
            broadcast_end,
        )
        fallback_df = filter_homeshopping_schedule(
            fallback_schedules,
            selected_homeshopping_channels,
            broadcast_start,
            broadcast_end,
        )
        if not fallback_df.empty:
            homeshopping_df = fallback_df
            homeshopping_errors = []
        elif fallback_errors and not homeshopping_errors:
            homeshopping_errors = fallback_errors

    return tv_df, homeshopping_df, tv_programs, tv_errors, homeshopping_errors


def render_input_form() -> tuple[bool, dict[str, Any]]:
    """상단 입력 영역을 렌더링하고 입력값을 반환합니다."""
    with st.form("search_form"):
        st.subheader("조회 조건")

        col1, col2, col3 = st.columns(3)
        broadcast_date = col1.date_input("방송날짜", value=date.today())
        start_time_text = col2.text_input("내 방송 시작시간", value="21:00")
        duration_minutes = col3.number_input("방송분", min_value=1, max_value=600, value=70, step=5)

        st.markdown("**TV 채널 선택**")
        tv_defaults = {
            "KBS1": True,
            "KBS2": True,
            "MBC": True,
            "SBS": True,
            "tvN": False,
            "종편": False,
        }
        tv_cols = st.columns(len(tv_defaults))
        selected_tv_options: list[str] = []
        for index, (label, checked) in enumerate(tv_defaults.items()):
            if tv_cols[index].checkbox(label, value=checked):
                selected_tv_options.append(label)

        st.markdown("**홈쇼핑 채널 선택**")
        hs_cols = st.columns(4)
        selected_homeshopping_channels: list[str] = []
        for index, channel_name in enumerate(HOMESHOPPING_CHANNELS):
            if hs_cols[index % 4].checkbox(channel_name, value=True):
                selected_homeshopping_channels.append(channel_name)

        submitted = st.form_submit_button("조회하기", type="primary")

    inputs = {
        "broadcast_date": broadcast_date,
        "start_time_text": start_time_text,
        "duration_minutes": int(duration_minutes),
        "selected_tv_options": selected_tv_options,
        "selected_homeshopping_channels": selected_homeshopping_channels,
    }
    return submitted, inputs


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.title(APP_TITLE)

    submitted, inputs = render_input_form()

    if "last_result" in st.session_state and "tv_programs" not in st.session_state["last_result"]:
        del st.session_state["last_result"]

    # 첫 화면에서도 기본값으로 한 번 조회되게 하여 초보자가 바로 결과 형태를 볼 수 있게 합니다.
    should_search = submitted or "last_result" not in st.session_state

    if should_search:
        if not inputs["selected_tv_options"]:
            st.error("TV 채널을 1개 이상 선택해주세요.")
            return

        if not inputs["selected_homeshopping_channels"]:
            st.error("홈쇼핑 채널을 1개 이상 선택해주세요.")
            return

        try:
            broadcast_start, broadcast_end = get_user_broadcast_window(
                inputs["broadcast_date"],
                inputs["start_time_text"],
                inputs["duration_minutes"],
            )
        except ValueError as exc:
            st.error(str(exc))
            return

        with st.spinner("편성표를 조회하는 중입니다."):
            tv_df, homeshopping_df, tv_programs, tv_errors, homeshopping_errors = run_search(
                inputs["selected_tv_options"],
                inputs["selected_homeshopping_channels"],
                broadcast_start,
                broadcast_end,
            )

        st.session_state["last_result"] = {
            "broadcast_start": broadcast_start,
            "broadcast_end": broadcast_end,
            "tv_df": tv_df,
            "homeshopping_df": homeshopping_df,
            "tv_programs": tv_programs,
            "tv_errors": tv_errors,
            "homeshopping_errors": homeshopping_errors,
        }

    if "last_result" in st.session_state:
        render_results(**st.session_state["last_result"])


if __name__ == "__main__":
    main()
